import numpy as np
import pandas as pd
import openpyxl
from .arbin_cycler import ArbinStep, ArbinCycle, ArbinCell
from .gamry_eis import EisSweep, EisCycle, EisCell

class CellBuilder:
    """
    This object builds the cell data structure from the raw Arbin data files.
    """

    def __init__(self) -> None:
        self._reload_data = False

    def read_B6_csv_data(self, cell: ArbinCell, file_path: str, steps: dict[str, list[int]], verbose: bool=False):
        """
        Read raw Arbin test data from a csv file such that only the data according to the steps
        in `steps` is read and stored.

        Parameters
        ----------
        cell : ArbinCell
            Cell that the data is to be loaded to.
        file_path : str
            Path to an Arbin raw data file according to batch B6. Must be a .csv file.
        steps : dict
            Keys are the step type (see ArbinStep.__init__) and values are lists of step numbers to be read
            and stored.
        """

        all_step_numbers = []
        for number_list in steps.values():
            for number in number_list:
                all_step_numbers.append(number)

        self.current_cycle_index = 0
        self.current_step_index = 0
        self.current_step_type = ''
        with open(file_path, 'rb') as data_file:

            # if cell has cycles in it take the last cycle index and step
            # index to be the current ones. This insures that cycles and steps
            # spread over multiple csv files are put in the correct cycle and step
            # in these objects.
            if cell.cycles:
                self.current_cycle_index = cell.cycles[-1].cycle_index
                if cell.cycles[-1].steps:
                    self.current_step_index = cell.cycles[-1].steps[-1].step_index

            is_first_line = True
            for line in data_file:
                # had to read the csv file as binary and then decode because of non ascii 
                # characters in the temperature header of Arbin files.
                line = line.decode('unicode_escape').strip()
                if is_first_line:
                    headers = line.split(',')
                    headers = self._fix_temperature_header(headers)
                    cell.update_headers(headers)
                    is_first_line = False
                    continue

                data = line.split(',')
                self._read_row(cell, data, all_step_numbers, steps, verbose)

            self.current_cycle_index = 0
            self.current_step_index = 0
            self.current_step_type = ''

    def read_leaf_characterization_excel_data(
        self, 
        file_path: str, 
        steps: dict[str, list[int]],
        verbose: bool=False
        ) -> list[ArbinCell]:
        """
        Load Nissan Leaf initial characterization data from the init_gen1_packX excel files.
        
        Parameters
        ----------
        cell : ArbinCell
            Cell that the data is to be loaded to.
        file_path : str
            Path to an Arbin raw data file according to batch B6. Must be a .csv file.
        steps : dict
            Keys are the step type (see ArbinStep.__init__) and values are lists of step numbers to be read
            and stored.
        """
        all_step_numbers = []
        for number_list in steps.values():
            for number in number_list:
                all_step_numbers.append(number)

        self.current_cycle_index = 0
        self.current_step_index = 0
        self.current_step_type = ''

        workbook = openpyxl.load_workbook(file_path, read_only=True)
        sheet_names = workbook.sheetnames[1:]
        channel_numbers = [int(sheet_name.split('_')[1]) for sheet_name in sheet_names]

        arbin_cells = []
        for idx, sheet_name in enumerate(sheet_names):
            arbin_cells.append(ArbinCell(channel_numbers[idx], channel_numbers[idx]))
            arbin_cell = arbin_cells[idx]
            sheet = workbook[sheet_name]
            headers = []
            for row_number, row in enumerate(sheet.iter_rows()):
                column_data = []
                for idx, column in enumerate(row):
                    if row_number == 0:
                        headers.append(column.value)
                    else:
                        column_data.append(column.value)
                if row_number == 0:
                    headers = self._fix_temperature_header(headers)
                    arbin_cell.update_headers(headers)
                else:   
                    self._read_row(arbin_cell, column_data, all_step_numbers, steps, verbose)
            
            self.current_cycle_index = 0
            self.current_step_index = 0
            self.current_step_type = ''
        
        return arbin_cells
    
    def merge_B6_eis_data(self, eis_cells: list[EisCell], arbin_cells: list[ArbinCell]) -> None:
        """
        Merge eis cells into the arbin cells. Ensures that eis steps are input into the correct ArbinCyles.

        Eis steps will only be merged into the Arbin objects if their is a correspoinding ArbinCell with a
        matching channel number and cycle number.

        Parameters
        ----------
        eis_cells : list[EisCell]
            List of EisCell to have steps added to the ArbinCells.
        arbin_cells : list[ArbinCell]
            List of ArbinCell to have eis steps added to.
        """
        eis_cell_dict = {cell.cell_number: cell for cell in eis_cells}
        eis_channel_dict = {cell_number: [cycle.cycle_index for cycle in cell] 
                            for cell_number, cell in eis_cell_dict.items()}

        for arbin_cell in arbin_cells:
            if arbin_cell.cell_number in eis_cell_dict.keys():
                eis_cycle_numbers = eis_channel_dict[arbin_cell.cell_number]
                for arbin_cycle in arbin_cell:
                    if arbin_cycle.cycle_index in eis_cycle_numbers:
                        eis_cell = eis_cell_dict[arbin_cell.cell_number]
                        for cycle in eis_cell:
                            if cycle.cycle_index == arbin_cycle.cycle_index:
                                for step in cycle:
                                    arbin_cycle.add_step(step)

                        


    def _read_row(
        self,
        cell: ArbinCell, 
        column_data: list, 
        all_step_numbers: list[int],
        steps: dict[str, list[int]],
        verbose: bool
        ) -> None:

        if cell.cycles:
            current_cycle = cell.cycles[-1]
            if cell.cycles[-1].steps:
                current_step = cell.cycles[-1].steps[-1]

        STEP_INDEX = int(column_data[3])
        CYCLE_INDEX = int(column_data[4])
        if CYCLE_INDEX != self.current_cycle_index:
            self.current_step_index = 0
            self.current_cycle_index = CYCLE_INDEX
            current_cycle = ArbinCycle(self.current_cycle_index)
            cell.add_cycle(current_cycle)
            if verbose:
                print(f'Processing test cycle {self.current_cycle_index}')
        if STEP_INDEX in all_step_numbers and STEP_INDEX != self.current_step_index:
            self.current_step_index = STEP_INDEX
            self.current_step_type = self._get_step_type(self.current_step_index, steps)
            current_step = ArbinStep(self.current_step_index, self.current_step_type)
            self._add_data_to_new_step(current_step, cell.headers, column_data)
            current_cycle.add_step(current_step)
        elif STEP_INDEX in all_step_numbers and STEP_INDEX == self.current_step_index:
            self._add_data_to_current_step(current_step, cell.headers, column_data)
        elif STEP_INDEX not in all_step_numbers:
            self.current_step_index = 0

    def _fix_temperature_header(self, headers: list[str]) -> list:
        """
        Fix the temperature headers so that the non-ascii symbols used to make the degree symbol
        are removed as they are different for different programs.

        Parameters
        ----------
        headers : list
            List of strings that make up the headers in an Arbin raw data file.
        """
        for idx, header in enumerate(headers):
            if header.startswith('Aux') and header.endswith('_1'):
                headers[idx] = 'Battery_Temperature(C)'
            if header.startswith('Aux') and header.endswith('_2'):
                headers[idx] = 'Chamber_Temperature(C)'

        return headers

    def _get_step_type(self, step_number: int, steps: dict[str, list[int]]) -> str:
        for step_type, step_numbers in steps.items():
            if step_number in step_numbers:
                return step_type

    def _add_data_to_new_step(self, step: ArbinStep, keys: list[str], values: list[str]) -> None:
        """
        come back and fix so that values are not all str.
        """
        for idx, header in enumerate(keys):
            step[header] = [values[idx]]

    def _add_data_to_current_step(self, step: ArbinStep, keys: list[str], values: list[str]) -> None:
        """
        come back and fix so that values are not all str.
        """
        for idx, header in enumerate(keys):
            step[header].append(values[idx])